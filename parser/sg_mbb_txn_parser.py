import pandas as pd
import re
import openpyxl
from openpyxl.utils import get_column_letter
import os
from pathlib import Path
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def find_transaction_description_column(df):
    """Find the column containing descriptions.
    
    First looks specifically for 'Description', then tries generic matches.
    """
    # First try the exact column name
    if "Description" in df.columns:
        return "Description"
    
    # Fall back to more generic search
    for col in df.columns:
        if any(term in str(col).lower() for term in ['description', 'particulars', 'details']):
            return col
    
    return None

def clean_customer_name(name):
    """Clean and standardize customer names, returning both clean name and extra info for description."""
    if not name:
        return '', ''
    
    # Convert to string and strip whitespace
    name = str(name).strip()
    extra_info = ''
    
    # Remove trailing periods
    name = name.rstrip('.')
    
    # Fix "SDN. BHD." to "SDN BHD"
    name = re.sub(r'SDN\.\s*BHD\.?', 'SDN BHD', name)
    
    # Fix abbreviated names
    if name.endswith('SB'):
        name = name.replace('SB', '').strip()

    # Extract numeric sequences and what follows them
    match = re.search(r'\s+(\d{8,}.*$)', name)
    if match:
        extra_info = match.group(1).strip()
        name = name[:match.start()].strip()
    
    # Extract invoice/document numbers
    match = re.search(r'\s+([A-Z]{2,3}[-\d]+.*$)', name)
    if match:
        extra_info = (extra_info + ' ' + match.group(1)).strip()
        name = name[:match.start()].strip()
    
    # Extract dates
    match = re.search(r'\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{2}.*$', name, flags=re.IGNORECASE)
    if match:
        extra_info = (extra_info + ' ' + match.group(0)).strip()
        name = name[:match.start()].strip()
    
    # Extract year
    match = re.search(r'\s+(20\d{2}).*$', name)
    if match:
        extra_info = (extra_info + ' ' + match.group(1)).strip()
        name = name[:match.start()].strip()
    
    # Extract repeated company name at the end
    words = name.split()
    if len(words) > 3:
        company_name = ' '.join(words[:3]).lower()
        remaining = ' '.join(words[3:]).lower()
        if remaining.startswith(company_name):
            extra_info = (extra_info + ' ' + ' '.join(words[3:])).strip()
            name = ' '.join(words[:3])
    
    return name.strip(), extra_info.strip()

def process_inward_fast_transaction(txn_desc):
    """Process Inward FAST transaction descriptions."""
    customer_name = ''
    description = ''
    
    # Remove the prefix "Inward FAST - "
    remaining = txn_desc[len("Inward FAST - "):]
    
    # Split by comma
    parts = remaining.split(", ", 1)  # Split by first comma only
    
    # First part is the customer name
    customer_name = parts[0]
    
    # Remove all periods from customer name
    customer_name = customer_name.replace('.', '')
    
    # Normalize PTE LTD format
    customer_name = re.sub(r'PTE\s*LTD', 'PTE LTD', customer_name)
    
    # Second part, if exists, is the description
    if len(parts) > 1:
        description = parts[1].strip()
        
        # No need to handle "OTHR-Other" specially anymore, keep as is
    
    # Final cleanup for customer names
    if customer_name.endswith(" PTE"):
        customer_name += " LTD"
    
    return customer_name, description

def process_inward_paynow_transaction(txn_desc):
    """Process Inward PayNow transaction descriptions."""
    customer_name = ''
    description = ''
    
    # Remove the prefix "Inward PayNow from "
    remaining = txn_desc[len("Inward PayNow from "):]
    
    # Look for code patterns that indicate a description
    patterns = ["BEXP-", "IVPT-", "OTHR-", "GDDS-", "SUPP-"]
    desc_start_idx = -1
    
    for pattern in patterns:
        idx = remaining.find(pattern)
        if idx != -1:
            desc_start_idx = idx
            break
    
    if desc_start_idx != -1:
        # Extract customer name and description
        customer_name = remaining[:desc_start_idx].strip()
        description = remaining[desc_start_idx:].strip()
    else:
        # Check for other separators like spaces followed by keywords
        for keyword in ["February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]:
            pattern = f" {keyword}"
            idx = remaining.find(pattern)
            if idx != -1:
                desc_start_idx = idx
                customer_name = remaining[:desc_start_idx].strip()
                description = remaining[desc_start_idx:].strip()
                break
        
        # If still no description found, the entire remaining is the customer name
        if desc_start_idx == -1:
            customer_name = remaining.strip()
    
    # Remove all periods from customer name
    customer_name = customer_name.replace('.', '')
    
    # Normalize PTE LTD format
    customer_name = re.sub(r'PTE\s*LTD', 'PTE LTD', customer_name)
    
    # Special case: do not modify "MOOD COLLECTIVES PTE" by adding LTD
    special_cases = ['MOOD COLLECTIVES PTE']
    if customer_name.endswith(" PTE") and customer_name not in special_cases:
        customer_name += " LTD"
    
    return customer_name, description

def process_giro_credit_transaction(txn_desc):
    """Process Giro Credit transaction descriptions."""
    customer_name = ''
    description = ''
    
    # Remove the prefix "Giro Credit from "
    remaining = txn_desc[len("Giro Credit from "):]
    
    # Look for code patterns that indicate a description
    patterns = ["BEXP-", "IVPT-", "OTHR-", "GDDS-", "SUPP-"]
    desc_start_idx = -1
    
    for pattern in patterns:
        idx = remaining.find(pattern)
        if idx != -1:
            desc_start_idx = idx
            break
    
    if desc_start_idx != -1:
        # Extract customer name and description
        customer_name = remaining[:desc_start_idx].strip()
        description = remaining[desc_start_idx:].strip()
    else:
        # No description found, the entire remaining is the customer name
        customer_name = remaining.strip()
    
    # Remove all periods from customer name
    customer_name = customer_name.replace('.', '')
    
    # Final cleanup for customer names
    if customer_name.endswith(" PTE"):
        customer_name += " LTD"
    
    # Normalize PTE LTD format
    customer_name = re.sub(r'PTE\s*LTD', 'PTE LTD', customer_name)
    
    return customer_name, description

def process_ib_transfer_transaction(txn_desc):
    """Process IB Transfer transaction descriptions."""
    # Remove the prefix "IB Transfer from "
    customer_name = txn_desc[len("IB Transfer from "):].strip()
    description = ''
    
    # Remove all periods from customer name
    customer_name = customer_name.replace('.', '')
    
    # Check for special case where the name includes a P at the end (like D'ZANDER INTERIORS P)
    if customer_name.endswith(" P"):
        # This should likely be PTE LTD but we'll keep it as is per the example
        pass
    # Final cleanup for customer names
    elif customer_name.endswith(" PTE"):
        customer_name += " LTD"
    
    # Normalize PTE LTD format
    customer_name = re.sub(r'PTE\s*LTD', 'PTE LTD', customer_name)
    
    return customer_name, description

def clean_description(description):
    """Clean transaction descriptions."""
    if not description:
        return ""
    
    description = str(description).strip()
    
    # Don't uppercase descriptions anymore - keep as is
    return description

def extract_transaction_info(txn_desc):
    """Extract customer name and description from transaction description."""
    if not txn_desc or pd.isna(txn_desc):
        return '', ''
    
    txn_desc = str(txn_desc).strip()
    
    # Process transactions with separate functions
    if txn_desc.startswith("Inward FAST - "):
        customer_name, description = process_inward_fast_transaction(txn_desc)
    elif txn_desc.startswith("Inward PayNow from "):
        customer_name, description = process_inward_paynow_transaction(txn_desc)
    elif txn_desc.startswith("Giro Credit from "):
        customer_name, description = process_giro_credit_transaction(txn_desc)
    elif txn_desc.startswith("IB Transfer from "):
        customer_name, description = process_ib_transfer_transaction(txn_desc)
    else:
        return '', ''
    
    # Clean up the extracted data but no longer get extra info - simpler approach
    customer_name = customer_name.strip()
    
    # No longer uppercase description - keep as is in the original format
    description = clean_description(description)
    
    return customer_name, description

def save_with_formulas(input_file, df, output_file):
    """Save DataFrame to Excel while preserving formulas in the original file."""
    try:
        # Load original workbook with formulas
        wb = openpyxl.load_workbook(input_file, data_only=False, keep_links=False)
        ws = wb.active
        
        # First determine if CUSTOMER_NAME and DESCRIPTION columns already exist
        header_row = ws[1]  # First row is the header
        header_values = {cell.value: cell.column for cell in header_row if cell.value}
        
        # Check if columns exist
        customer_col = None
        desc_col = None
        
        for col_name, col_idx in header_values.items():
            if col_name == "CUSTOMER_NAME":
                customer_col = col_idx
            elif col_name == "DESCRIPTION":
                desc_col = col_idx
        
        # If columns don't exist yet, create them right after the Transaction Description column
        if not customer_col:
            customer_col = ws.max_column + 1
            ws.cell(row=1, column=customer_col).value = "CUSTOMER_NAME"
            
            if not desc_col:
                desc_col = ws.max_column + 1
                ws.cell(row=1, column=desc_col).value = "DESCRIPTION"
        
        # If only DESCRIPTION column is missing, add it after CUSTOMER_NAME
        elif not desc_col:
            next_col = customer_col + 1
            ws.insert_cols(next_col)
            ws.cell(row=1, column=next_col).value = "DESCRIPTION"
            desc_col = next_col
        
        # Create a mapping of row indices between DataFrame and Excel
        # Excel rows start at 1, with row 1 being the header
        df_to_excel_row_map = {}
        
        # If there's a transaction date column in both Excel and DataFrame, use it to map rows
        date_col_excel = None
        date_col_df = None
        
        # Look for date column in Excel
        for col_name, col_idx in header_values.items():
            if "Date" in str(col_name):
                date_col_excel = col_idx
                break
        
        # Look for date column in DataFrame
        for col in df.columns:
            if "Date" in str(col):
                date_col_df = col
                break
        
        # If we found date columns in both, use them to map rows
        if date_col_excel and date_col_df:
            print(f"Using date column for row mapping: {date_col_df}")
            
            # Create a dictionary of date values to row indices in Excel
            excel_dates = {}
            for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                cell_value = ws.cell(row=row_idx, column=date_col_excel).value
                if cell_value:
                    # Handle potential datetime objects
                    if hasattr(cell_value, 'strftime'):
                        cell_value = cell_value.strftime('%Y-%m-%d %H:%M:%S')
                    excel_dates[str(cell_value)] = row_idx
            
            # Map DataFrame rows to Excel rows using dates
            for df_idx, row in df.iterrows():
                date_value = row[date_col_df]
                if date_value is not None:
                    # Handle potential datetime objects
                    if hasattr(date_value, 'strftime'):
                        date_value = date_value.strftime('%Y-%m-%d %H:%M:%S')
                    
                    # Look up the Excel row for this date
                    if str(date_value) in excel_dates:
                        df_to_excel_row_map[df_idx] = excel_dates[str(date_value)]
        
        # If no date mapping was possible, use simple row-to-row mapping
        if not df_to_excel_row_map:
            print("Using simple row-to-row mapping")
            # Just map DataFrame rows directly to Excel rows (2+)
            for df_idx in range(len(df)):
                df_to_excel_row_map[df_idx] = df_idx + 2  # +2 because Excel rows start at 1 and row 1 is header
        
        # Now update the data for each row using the mapping
        for df_idx, row in df.iterrows():
            if df_idx in df_to_excel_row_map:
                excel_row = df_to_excel_row_map[df_idx]
                
                if pd.notna(row["CUSTOMER_NAME"]) and row["CUSTOMER_NAME"]:
                    ws.cell(row=excel_row, column=customer_col).value = row["CUSTOMER_NAME"]
                
                if pd.notna(row["DESCRIPTION"]) and row["DESCRIPTION"]:
                    ws.cell(row=excel_row, column=desc_col).value = row["DESCRIPTION"]
        
        # Save workbook
        wb.save(output_file)
        print(f"Saved to {output_file} with formulas preserved")
        print(f"CUSTOMER_NAME is in column {get_column_letter(customer_col)}")
        print(f"DESCRIPTION is in column {get_column_letter(desc_col)}")
        return True
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False

def process_transactions(input_file, output_file):
    """
    Main function to process bank transactions and extract customer names and descriptions.
    
    Parameters:
    input_file (str): Path to the input Excel file
    output_file (str): Path to save the output Excel file
    
    Returns:
    bool: True if processing was successful, False otherwise
    """
    try:
        print(f"Processing file: {input_file}")
        
        # Load the workbook directly for processing
        wb = openpyxl.load_workbook(input_file, data_only=False, keep_links=False)
        ws = wb.active
        
        # Find description column in Excel
        txn_desc_col = None
        header_row = ws[1]  # First row is the header
        
        for cell in header_row:
            if cell.value == "Description":
                txn_desc_col = cell.column
                break
        
        if not txn_desc_col:
            print("Error: Could not find Description column")
            return False
        
        print(f"Found Description column at position {get_column_letter(txn_desc_col)}")
        
        # Find or create CUSTOMER_NAME and DESCRIPTION columns
        customer_col = None
        desc_col = None
        
        for cell in header_row:
            if cell.value == "CUSTOMER_NAME":
                customer_col = cell.column
            elif cell.value == "DESCRIPTION":
                desc_col = cell.column
        
        # Create columns if they don't exist
        if not customer_col:
            # Add column at the end
            next_col = ws.max_column + 1
            ws.cell(row=1, column=next_col).value = "CUSTOMER_NAME"
            customer_col = next_col
            print(f"Created CUSTOMER_NAME column at position {get_column_letter(customer_col)}")
        
        if not desc_col:
            # Insert column after CUSTOMER_NAME
            next_col = customer_col + 1
            ws.insert_cols(next_col)
            ws.cell(row=1, column=next_col).value = "DESCRIPTION"
            desc_col = next_col
            print(f"Created DESCRIPTION column at position {get_column_letter(desc_col)}")
        
        # Process each row directly in Excel
        customer_count = 0
        desc_count = 0
        total_rows = 0
        
        # Start from row 2 (skip header)
        for row_idx in range(2, ws.max_row + 1):
            # Get transaction description
            txn_desc = ws.cell(row=row_idx, column=txn_desc_col).value
            
            if txn_desc:
                total_rows += 1
                # Extract customer name and description
                customer_name, description = extract_transaction_info(txn_desc)
                
                # Update cells
                if customer_name:
                    ws.cell(row=row_idx, column=customer_col).value = customer_name
                    customer_count += 1
                
                if description:
                    ws.cell(row=row_idx, column=desc_col).value = description
                    desc_count += 1
        
        # Save the workbook
        wb.save(output_file)
        
        print(f"Processed {total_rows} transactions")
        print(f"Extracted {customer_count} customer names")
        print(f"Extracted {desc_count} descriptions")
        print(f"Saved to {output_file} with formulas preserved")
        
        return True
    
    except Exception as e:
        print(f"Error processing transactions: {e}")
        import traceback
        traceback.print_exc()
        return False

# Example usage
if __name__ == "__main__":
    input_file = Path('data/downloads') / 'new_rows_JOTEX PTE LTD MAYBANK SG 2025.csv'
    output_file = Path('data/temp') / 'SG_MBB_2025_processed.csv'
    
    success = process_transactions(input_file, output_file)
    
    if success:
        print("Processing completed successfully!")
    else:
        print("Processing failed. Check the error messages above.")